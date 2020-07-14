import numpy as np
from openpyxl import load_workbook


def readExcel(path=None, sheet_number=0, sheet_name=None, row_index=1, column_index=1, row_num=None, column_num=None,
              y_column=1):
    """

    :param sheet_number: excel中第几个sheet，默认是第一个sheet
    :param path:Excel文档的绝对路径
    :param sheet_name:数据所属sheet的名称,如果不想给sheet编号，就直接给name
    :param row_index: 读取数据的开始行数
    :param column_index: 读取数据的开始列数
    :param row_num:数据的行数
    :param column_num:数据的列数
    :param y_column: 总是有一列和其他的不一样，可以单独提取出来

    :return:
    """
    ''' 读取Ecxel中的数据

    参数：
        path: Excel文档的绝对路径
        sheet：数据所属sheet的名称
        row_num: 数据的列数
        column: 数据的行数
    返回：
        返回一个tuple（元组），得到data_X和data_Y
    注意：
        data_Y[i - 1] = float(sheet.cell(row=i, column=5).value)，此处数据Y的位置在列5，所以column为5，可更改
    '''
    book = load_workbook(filename=path)  # 按照路径读取Excel
    if sheet_name is not None:
        sheet = book[sheet_name]  # 读取Excel中的以名称指定的sheet
    else:
        sheet = book[book.sheetnames[sheet_number]]  # 读取Excel中的以序号指定sheet
    if row_num is not None and column_num is not None:
        row_num_read = row_num
        column_num_read = column_num
    elif row_num is not None and column_num is None:
        row_num_read = row_num
        column_num_read = sheet.max_column - column_index + 1
    elif row_num is None and column_num is not None:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = column_num
    else:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = sheet.max_column - column_index + 1

    data_X = np.empty((column_num_read, row_num_read))  # 用于存储数据的array，相较于np.ones只分配不初始化
    data_Y = np.empty(row_num_read)
    for j in range(column_index, column_index + column_num_read):
        for i in range(row_index, row_index + row_num_read):
            data_X[j - column_index][i - row_index] = np.abs(float(sheet.cell(row=i, column=j).value))
            if j == column_index:
                data_Y[i - row_index] = float(sheet.cell(row=i, column=y_column).value)
    return data_X, data_Y
