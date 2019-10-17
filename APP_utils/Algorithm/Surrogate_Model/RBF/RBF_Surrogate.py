import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def readExcel(path, sheet, row_num, column_num):
    # 按照路径读取Excel
    book = load_workbook(filename=path)
    # 读取Excel中的指定sheet
    sheet = book[sheet]
    # 用于存储数据的array，相较于np.ones只分配不初始化
    data_X = np.empty((row_num, column_num))
    data_Y = np.empty(row_num)
    for i in range(1, row_num + 1):
        for j in range(1, column_num + 1):
            data_X[i - 1][j - 1] = float(sheet.cell(row=i, column=j).value)
            if j == 1:
                data_Y[i - 1] = float(sheet.cell(row=i, column=5).value)
    return data_X, data_Y


def Gaussian(x, c, s):
    return np.exp(-(x - c) ** 2 / (2 * s ** 2))


def Multiquadric(x, c, s):
    return np.sqrt((x - c) ** 2 + s ** 2)


class RBF(object):
    # def RBF(X, Y, X_Pre):
    def __init__(self, rbf=Multiquadric, std=0, w=0, x=np.array([])):
        self.rbf = rbf  # rbf名称
        self.std = std
        self.w = w
        self.x = x

    def fit(self, X, Y):
        list_Multiquadric_result = []
        self.x = X
        # self.std = (max(X) - min(X)) / len(X)
        max_distance_lsit = []
        for i in range(len(X)):
            max_distance_lsit.append(max([np.linalg.norm(m) for m in X - np.array([X[i]])]))
        self.std = max(max_distance_lsit) / len(X)

        for i in range(X.shape[0]):
            list_temp = []
            for j in range(X.shape[0]):
                list_temp.append(self.rbf(X[i], X[j], self.std))
            list_Multiquadric_result.append(np.array(list_temp).ravel())
        Gaussian_result = np.array(list_Multiquadric_result)
        self.w = np.linalg.pinv(Gaussian_result).dot(Y)
        print(self.w)
        return self.w

    def predict(self, X_Pre):
        list_pre_x = []
        for i in range(X_Pre.shape[0]):
            list_temp = []
            for j in range(self.x.shape[0]):
                list_temp.append(self.rbf(X_Pre[i], self.x[j], self.std))
            list_pre_x.append(np.array(list_temp).ravel())
        Y_Pre = np.array(list_pre_x).dot(self.w)
        return Y_Pre


if __name__ == "__main__":
    path_excel = r"C:\Users\asus\Desktop\History\History_codes\NewPython\APP_utils\Algorithm\data\test7fun.xlsx"
    data_real = readExcel(path_excel, "Sheet1", 30, 3)
    data_pre = readExcel(path_excel, "Sheet2", 30, 3)
    d = np.array([-17, -13, -9, -5, -1, 0, 1, 5, 9, 13, 17])
    y = np.array([22.3, 16.85, 11.4, 5.9501, 0.95417, 0.5, 0.95417, 5.9501, 11.4, 16.85, 22.3])
    d_pred = np.arange(-17, 18)
    # Y_pre = RBF()
    # Y_pre.fit(data_real[0], data_real[1])
    # y_Pre = Y_pre.predict(data_pre[0])
    # RR = 1 - (np.sum(np.square(data_pre[1] - y_Pre)) / np.sum(np.square(data_pre[1] - np.mean(data_pre[1]))))
    # print(RR)
    Y_pre1 = RBF()
    Y_pre1.fit(d, y)
    y_Pre1 = Y_pre1.predict(d_pred)
    # plt.plot(d, y, color='#ff0000', marker='+', linestyle='-',
    #          label='z-real')
    # plt.plot(data_pre[0], y_Pre, color='#0000ff', marker='+', linestyle='-.',
    #          label='z-predict')

    plt.plot(d, y, color='#ff0000', marker='+', linestyle='-',
             label='z-real')
    plt.plot(d_pred, y_Pre1, color='#0000ff', marker='+', linestyle='-.',
             label='z-predict')
    # RR = 1 - (np.sum(np.square(y - y_Pre1)) / np.sum(np.square(y - np.mean(y))))
    # print(RR)
    plt.legend()
    plt.show()
