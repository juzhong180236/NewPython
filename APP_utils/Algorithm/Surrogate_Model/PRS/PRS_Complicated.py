import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def readExcel(path, sheet, row_num, column_num):
    # 按照路径读取Excel
    book = load_workbook(filename=path)
    # 读取Excel中的指定sheet
    sheet = book[sheet]
    # 用于存储数据的array，相较于np.ones只分配不初始化
    data_X = np.empty((row_num, column_num))
    data_Y = np.empty(row_num)
    for i in range(1, row_num + 1):
        for j in range(1, column_num + 1):
            data_X[i - 1][j - 1] = float(sheet.cell(row=i, column=j).value)
            if j == 1:
                data_Y[i - 1] = float(sheet.cell(row=i, column=5).value)
    return data_X, data_Y


class PRS(object):
    def __init__(self, m=7, w=0):
        self.m = m
        self.w = w

    def fit(self, X, Y):
        # 把训练点的x输入
        # data_PRS_result = np.empty(shape=(X.shape[0], self.m + 1), dtype=object)
        list_PRS_result = []
        for i in range(X.shape[0]):  # m
            list_result = X[i]
            list_combine = X[i]
            for j in range(self.m):
                list_temp = [m * list_result for m in np.array(X[i])]
                list_result = np.array(list_temp).flatten()
                list_combine = np.concatenate((list_combine, list_result))
                # data_PRS_result[i][j] = X[i] ** j
            list_PRS_result.append(list_combine)
        PRS_result = np.array(list_PRS_result)

        # 根据Y和伪逆求出w
        self.w = np.linalg.pinv(PRS_result).dot(Y)
        return self.w

    def predict(self, X_Pre):
        list_pre_x = []
        for i in range(X_Pre.shape[0]):
            list_result = X_Pre[i]
            list_combine = X_Pre[i]
            for j in range(self.m):
                list_temp = [m * list_result for m in np.array(X_Pre[i])]
                list_result = np.array(list_temp).flatten()
                list_combine = np.concatenate((list_combine, list_result))
            list_pre_x.append(list_combine)
        Y_Pre = np.array(list_pre_x).dot(self.w)
        return Y_Pre


if __name__ == "__main__":
    path_excel = r"C:\Users\asus\Desktop\History\History_codes\NewPython\APP_utils\Algorithm\data\test7fun.xlsx"
    data_real = readExcel(path_excel, "Sheet1", 30, 3)
    data_pre = readExcel(path_excel, "Sheet2", 30, 3)
    d = np.array([-17, -13, -9, -5, -1, 0, 1, 5, 9, 13, 17])
    y = np.array([22.3, 16.85, 11.4, 5.9501, 0.95417, 0.5, 0.95417, 5.9501, 11.4, 16.85, 22.3])
    d_pred = np.arange(-17, 18)

    prs1 = PRS(m=3)
    prs1.fit(data_real[0], data_real[1])
    Y_pre = prs1.predict(data_pre[0])
    RR = 1 - (np.sum(np.square(data_pre[1] - Y_pre)) / np.sum(np.square(data_pre[1] - np.mean(data_pre[1]))))
    print(RR)
    # prs2 = PRS(m=7)
    # prs2.fit(d, y)
    # Y_pre1 = prs2.predict(d_pred)

    # plt.plot(data_real[0], data_real[1], color='#ff0000', marker='+', linestyle='-',
    #          label='z-real')
    # plt.plot(data_real[0], Y_pre, color='#0000ff', marker='+', linestyle='-.',
    #          label='z-predict')
    # plt.plot(data_real[0], data_real[1], color='#ff0000', marker='+', linestyle='-',
    #     #          label='z-real')
    plt.plot(data_pre[0], data_pre[1], color='#ff0000', marker='+', linestyle='-',
             label='z-real')
    plt.plot(data_pre[0], Y_pre, color='#0000ff', marker='+', linestyle='-.',
             label='z-predict')
    # plt.plot(d, y, color='#ff0000', marker='+', linestyle='-',
    #          label='z-real')
    # plt.plot(d_pred, Y_pre1, color='#0000ff', marker='+', linestyle='-.',
    #          label='z-predict')
    # print((Y_pre - data_real[1]) / data_real[1] * 100 )

    plt.legend()
    plt.show()
