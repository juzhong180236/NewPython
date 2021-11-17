import numpy as np
from Demo.Crane_boom_2021.read_data import Read_Data
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
from Demo.Ansys_Data_Utils_2021.Surrogate_Models.RBF import RBF
from Demo.Ansys_Data_Utils_2021.Surrogate_Models.MFS_RBF import MFS_RBF
from scipy.stats import norm
from matplotlib.patches import ConnectionPatch
from openpyxl import load_workbook
import os


def readExcel(path=None, sheet_number=0, sheet_name=None, row_index=1, column_index=1, row_num=None, column_num=None,
              y_column=1):
    """

    :param sheet_number: excel中第几个sheet，默认是第一个sheet
    :param path:Excel文档的绝对路径
    :param sheet_name:数据所属sheet的名称,如果不想给sheet编号，就直接给name
    :param row_index: 读取数据的开始行数
    :param column_index: 读取数据的开始列数
    :param row_num:数据的行数
    :param column_num:数据的列数
    :param y_column: 总是有一列和其他的不一样，可以单独提取出来

    :return:
    """
    ''' 读取Ecxel中的数据

    参数：
        path: Excel文档的绝对路径
        sheet：数据所属sheet的名称
        row_num: 数据的列数
        column: 数据的行数
    返回：
        返回一个tuple（元组），得到data_X和data_Y
    注意：
        data_Y[i - 1] = float(sheet.cell(row=i, column=5).value)，此处数据Y的位置在列5，所以column为5，可更改
    '''
    book = load_workbook(filename=path)  # 按照路径读取Excel
    if sheet_name is not None:
        sheet = book[sheet_name]  # 读取Excel中的以名称指定的sheet
    else:
        sheet = book[book.sheetnames[sheet_number]]  # 读取Excel中的以序号指定sheet
    if row_num is not None and column_num is not None:
        row_num_read = row_num
        column_num_read = column_num
    elif row_num is not None and column_num is None:
        row_num_read = row_num
        column_num_read = sheet.max_column - column_index + 1
    elif row_num is None and column_num is not None:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = column_num
    else:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = sheet.max_column - column_index + 1

    data_X = np.empty((column_num_read, row_num_read))  # 用于存储数据的array，相较于np.ones只分配不初始化
    data_Y = np.empty(row_num_read)
    for j in range(column_index, column_index + column_num_read):
        for i in range(row_index, row_index + row_num_read):
            data_X[j - column_index][i - row_index] = np.abs(float(sheet.cell(row=i, column=j).value))
            if j == column_index:
                data_Y[i - row_index] = float(sheet.cell(row=i, column=y_column).value)
    return data_X, data_Y


def Read_TData(path, _row_index):
    data_y, data_x = readExcel(path, row_index=_row_index, column_index=4, y_column=1)
    return data_x, data_y


def Read_SData(path):
    isExisted = os.path.exists(path)
    if not isExisted:
        print(path)
        print('上面列出的路径不存在，请设置正确路径！')
        return
    else:
        print('目录[' + path + ']存在,正在读取...')

    list_different_angle = []

    file_content = open(path, 'rt')
    list_1 = []
    list_2 = []
    list_3 = []
    list_4 = []
    list_5 = []
    list_6 = []
    list_7 = []
    list_8 = []
    list_angle = []
    i = 0
    for line in file_content:
        i += 1
        if i % 16 == 1:
            list_1.append(np.abs(float(line)))
        elif i % 16 == 2:
            list_angle.append(float(line))
        elif i % 16 == 3:
            list_2.append(float(line))
        elif i % 16 == 5:
            list_3.append(float(line))
        elif i % 16 == 7:
            list_4.append(float(line))
        elif i % 16 == 9:
            list_5.append(float(line))
        elif i % 16 == 11:
            list_6.append(float(line))
        elif i % 16 == 13:
            list_7.append(float(line))
        elif i % 16 == 15:
            list_8.append(float(line))
        else:
            continue
    file_content.close()
    return list_1, list_2, list_3, list_4, list_5, list_6, list_7, list_8, list_angle


""" 
2021.11.05
起重机臂架论文第二次回复中的应力预测图，带置信区间
"""
""" 顔色 """
# cmaps = OrderedDict()
load_np_txt_path = r"D:\Alai\paper_Alai\【1】期刊论文\【1】Journal of Mechanical Design\起重机臂架论文\crane_boom\\"

"""
所谓low_fidelity，其实就是Beam188单元
"""
path_prefix = r"H:\Code\DT_Crane_Boom_v1.0\APP_models\\"
path_arr = \
    {
        "low": r"pre_low_fidelity_truss_point\stress_point_more_nodes\\",
        "high": r"pre_high_fidelity_truss_point\stress_point_less_samples\\",
        "verification": r"pre_verification_truss_point\stress_eighteen_nodes\\",
    }
rd = Read_Data()
""" 2020.12.21
读入高低保真的18个节点应力数据
"""
list_stress_low = rd.read_stress(path_prefix + path_arr["low"], mode='v')
list_stress_high = rd.read_stress(path_prefix + path_arr["high"], mode='v')
list_stress_verification = rd.read_stress(path_prefix + path_arr["verification"], mode='v')
""" 2020.12.21
每个点的应力放入一个数组，二维数组 
低保真样本18*10*10=18*100
高保真样本18*6
验证18*37*19=18*703
"""
array_real_stress_low = np.asarray(list_stress_low).T
array_real_stress_high = np.asarray(list_stress_high).T
_array_real_stress_verification = np.asarray(list_stress_verification).T

indices_degree = np.delete(np.arange(0, 37), np.arange(4, 36, 4))
indices_force = np.delete(np.arange(0, 19), np.arange(2, 18, 2))
list_real_stress_verification = []
for _d in _array_real_stress_verification:
    d = _d.reshape(19, 37)[indices_force, :]
    d1 = d[:, indices_degree]
    list_real_stress_verification.append(d1)
array_real_stress_verification = np.asarray(list_real_stress_verification)


def create_mf_rbf(_independent_variables_low, _independent_variables_high,
                  _dependent_variables_low, _dependent_variables_high):
    mf_rbf_stress_list = []

    low_model_stress_w_list = []
    stds = None
    rbf_type = 'mq'
    x_high = None
    bf_sigma = None
    omega_list = []

    for _i_point in range(len(_dependent_variables_low)):
        mf_rbf_stress = MFS_RBF()
        mf_rbf_stress.fit(_independent_variables_low,
                          _dependent_variables_low[_i_point].reshape(-1, 1),
                          _independent_variables_high,
                          _dependent_variables_high[_i_point].reshape(-1, 1),
                          bf_type="G",
                          )
        mf_rbf_stress_list.append(mf_rbf_stress)

        low_model_stress_w_list.append(mf_rbf_stress.low_model.w.tolist())
        if _i_point == 0:
            stds = mf_rbf_stress.low_model.std
            x_high = mf_rbf_stress.x_high.tolist()
            bf_sigma = mf_rbf_stress.bf_sigma.tolist()
        omega_list.append(mf_rbf_stress.omega.tolist())
    return mf_rbf_stress_list


def create_train_independent_variables(_force_arr, _degree_arr):
    _combine = []
    for _iForce in range(len(_force_arr)):
        for _iDegree in range(len(_degree_arr)):
            _combine.append((_force_arr[_iForce], _degree_arr[_iDegree]))
    _independent_variables = np.array(_combine)
    return _independent_variables


""" 训练集数据：自变量 """
# low_samples
force_arr_low = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]  # 5 10 15 20 25 30 35 40 45 50
degree_arr_low = [0, 8, 16, 24, 32, 40, 48, 56, 64, 72]  # 0 8 16 24 32 40 48 56 64 72
# high_samples
train_high = np.asarray([[5, 0], [5, 72], [20, 24], [35, 48], [50, 0], [50, 72]])
# verification_data
'''2020.12.21
这里一定要注意使用float，不然用整数的话kriging工作会出问题
'''
_load_arr = np.loadtxt(load_np_txt_path + r"dynamic_train\load_train_dynamic.txt")
# print("荷重传感器静止时均值与标准差：", _load.flatten().mean(), _load.flatten().std())
# load_std.append(_load.flatten().std())
# load_list.append(_load)
_angle_arr = np.loadtxt(load_np_txt_path + r"dynamic_train\angle_train_dynamic.txt")
# print("角度传感器静止时均值与标准差：", _angle.flatten().mean(), _angle.flatten().std())
# angle_std.append(_angle.flatten().std())
# angle_list.append(_angle)
# print("荷重传感器静止时2*sigma：", np.mean(load_std) * 2)
# print("角度传感器静止时2*sigma：", np.mean(angle_std) * 2)
independent_variables = []
for i in range(len(_load_arr)):
    _temp = []
    for j in range(len(_load_arr[0])):
        _temp.append([_load_arr[i][j], _angle_arr[i][j]])
    independent_variables.append(_temp)
independent_variables = np.array(independent_variables)
print(independent_variables.shape)
train_low = create_train_independent_variables(force_arr_low, degree_arr_low)
multirbf_list = create_mf_rbf(train_low, train_high, array_real_stress_low, array_real_stress_high)

# list_test_predict_stress_multi = []
# for i in range(len(multirbf_list)):
#     _temp_multi = []
#     for j in range(len(independent_variables)):
#         _temp_multi.append(multirbf_list[i].predict(independent_variables[j]))
#     list_test_predict_stress_multi.append(_temp_multi)
# test_predict_stress_multi = np.array(list_test_predict_stress_multi)
# print(test_predict_stress_multi[0][0])

list_test_predict_stress_multi = []
for j in range(len(independent_variables)):
    list_test_predict_stress_multi.append(multirbf_list[0].predict(independent_variables[j]).flatten())
test_predict_stress_multi = np.array(list_test_predict_stress_multi)
print(test_predict_stress_multi.shape)
# np.savetxt(load_np_txt_path + r"dynamic_train\stress_train_dynamic.txt", test_predict_stress_multi)
