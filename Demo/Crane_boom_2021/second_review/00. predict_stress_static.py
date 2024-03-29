import numpy as np
from Demo.Crane_boom_2021.read_data import Read_Data
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
from Demo.Ansys_Data_Utils_2021.Surrogate_Models.RBF import RBF
from Demo.Ansys_Data_Utils_2021.Surrogate_Models.MFS_RBF import MFS_RBF
from scipy.stats import norm
from matplotlib.patches import ConnectionPatch
from openpyxl import load_workbook
import os


def readExcel(path=None, sheet_number=0, sheet_name=None, row_index=1, column_index=1, row_num=None, column_num=None,
              y_column=1):
    """

    :param sheet_number: excel中第几个sheet，默认是第一个sheet
    :param path:Excel文档的绝对路径
    :param sheet_name:数据所属sheet的名称,如果不想给sheet编号，就直接给name
    :param row_index: 读取数据的开始行数
    :param column_index: 读取数据的开始列数
    :param row_num:数据的行数
    :param column_num:数据的列数
    :param y_column: 总是有一列和其他的不一样，可以单独提取出来

    :return:
    """
    ''' 读取Ecxel中的数据

    参数：
        path: Excel文档的绝对路径
        sheet：数据所属sheet的名称
        row_num: 数据的列数
        column: 数据的行数
    返回：
        返回一个tuple（元组），得到data_X和data_Y
    注意：
        data_Y[i - 1] = float(sheet.cell(row=i, column=5).value)，此处数据Y的位置在列5，所以column为5，可更改
    '''
    book = load_workbook(filename=path)  # 按照路径读取Excel
    if sheet_name is not None:
        sheet = book[sheet_name]  # 读取Excel中的以名称指定的sheet
    else:
        sheet = book[book.sheetnames[sheet_number]]  # 读取Excel中的以序号指定sheet
    if row_num is not None and column_num is not None:
        row_num_read = row_num
        column_num_read = column_num
    elif row_num is not None and column_num is None:
        row_num_read = row_num
        column_num_read = sheet.max_column - column_index + 1
    elif row_num is None and column_num is not None:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = column_num
    else:
        row_num_read = sheet.max_row - row_index + 1
        column_num_read = sheet.max_column - column_index + 1

    data_X = np.empty((column_num_read, row_num_read))  # 用于存储数据的array，相较于np.ones只分配不初始化
    data_Y = np.empty(row_num_read)
    for j in range(column_index, column_index + column_num_read):
        for i in range(row_index, row_index + row_num_read):
            data_X[j - column_index][i - row_index] = np.abs(float(sheet.cell(row=i, column=j).value))
            if j == column_index:
                data_Y[i - row_index] = float(sheet.cell(row=i, column=y_column).value)
    return data_X, data_Y


def Read_TData(path, _row_index):
    data_y, data_x = readExcel(path, row_index=_row_index, column_index=4, y_column=1)
    return data_x, data_y


def Read_SData(path):
    isExisted = os.path.exists(path)
    if not isExisted:
        print(path)
        print('上面列出的路径不存在，请设置正确路径！')
        return
    else:
        print('目录[' + path + ']存在,正在读取...')

    list_different_angle = []

    file_content = open(path, 'rt')
    list_1 = []
    list_2 = []
    list_3 = []
    list_4 = []
    list_5 = []
    list_6 = []
    list_7 = []
    list_8 = []
    list_angle = []
    i = 0
    for line in file_content:
        i += 1
        if i % 16 == 1:
            list_1.append(np.abs(float(line)))
        elif i % 16 == 2:
            list_angle.append(float(line))
        elif i % 16 == 3:
            list_2.append(float(line))
        elif i % 16 == 5:
            list_3.append(float(line))
        elif i % 16 == 7:
            list_4.append(float(line))
        elif i % 16 == 9:
            list_5.append(float(line))
        elif i % 16 == 11:
            list_6.append(float(line))
        elif i % 16 == 13:
            list_7.append(float(line))
        elif i % 16 == 15:
            list_8.append(float(line))
        else:
            continue
    file_content.close()
    return list_1, list_2, list_3, list_4, list_5, list_6, list_7, list_8, list_angle


""" 
2021.11.05
起重机臂架论文第二次回复中的应力预测图，带置信区间
"""
""" 顔色 """
# cmaps = OrderedDict()
load_np_txt_path = r"D:\Alai\paper_Alai\【1】期刊论文\【1】Journal of Mechanical Design\起重机臂架论文\crane_boom\\"

"""
所谓low_fidelity，其实就是Beam188单元
"""
path_prefix = r"H:\Code\DT_Crane_Boom_v1.0\APP_models\\"
path_arr = \
    {
        "low": r"pre_low_fidelity_truss_point\stress_point_more_nodes\\",
        "high": r"pre_high_fidelity_truss_point\stress_point_less_samples\\",
        "verification": r"pre_verification_truss_point\stress_eighteen_nodes\\",
    }
rd = Read_Data()
""" 2020.12.21
读入高低保真的18个节点应力数据
"""
list_stress_low = rd.read_stress(path_prefix + path_arr["low"], mode='v')
list_stress_high = rd.read_stress(path_prefix + path_arr["high"], mode='v')
list_stress_verification = rd.read_stress(path_prefix + path_arr["verification"], mode='v')
""" 2020.12.21
每个点的应力放入一个数组，二维数组 
低保真样本18*10*10=18*100
高保真样本18*6
验证18*37*19=18*703
"""
array_real_stress_low = np.asarray(list_stress_low).T
array_real_stress_high = np.asarray(list_stress_high).T
_array_real_stress_verification = np.asarray(list_stress_verification).T

indices_degree = np.delete(np.arange(0, 37), np.arange(4, 36, 4))
indices_force = np.delete(np.arange(0, 19), np.arange(2, 18, 2))
list_real_stress_verification = []
for _d in _array_real_stress_verification:
    d = _d.reshape(19, 37)[indices_force, :]
    d1 = d[:, indices_degree]
    list_real_stress_verification.append(d1)
array_real_stress_verification = np.asarray(list_real_stress_verification)


def create_mf_rbf(_independent_variables_low, _independent_variables_high,
                  _dependent_variables_low, _dependent_variables_high):
    mf_rbf_stress_list = []

    low_model_stress_w_list = []
    stds = None
    rbf_type = 'mq'
    x_high = None
    bf_sigma = None
    omega_list = []

    for _i_point in range(len(_dependent_variables_low)):
        mf_rbf_stress = MFS_RBF()
        mf_rbf_stress.fit(_independent_variables_low,
                          _dependent_variables_low[_i_point].reshape(-1, 1),
                          _independent_variables_high,
                          _dependent_variables_high[_i_point].reshape(-1, 1),
                          bf_type="G",
                          )
        mf_rbf_stress_list.append(mf_rbf_stress)

        low_model_stress_w_list.append(mf_rbf_stress.low_model.w.tolist())
        if _i_point == 0:
            stds = mf_rbf_stress.low_model.std
            x_high = mf_rbf_stress.x_high.tolist()
            bf_sigma = mf_rbf_stress.bf_sigma.tolist()
        omega_list.append(mf_rbf_stress.omega.tolist())
    return mf_rbf_stress_list


def create_train_independent_variables(_force_arr, _degree_arr):
    _combine = []
    for _iForce in range(len(_force_arr)):
        for _iDegree in range(len(_degree_arr)):
            _combine.append((_force_arr[_iForce], _degree_arr[_iDegree]))
    _independent_variables = np.array(_combine)
    return _independent_variables


""" 训练集数据：自变量 """
# low_samples
force_arr_low = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]  # 5 10 15 20 25 30 35 40 45 50
degree_arr_low = [0, 8, 16, 24, 32, 40, 48, 56, 64, 72]  # 0 8 16 24 32 40 48 56 64 72
# high_samples
train_high = np.asarray([[5, 0], [5, 72], [20, 24], [35, 48], [50, 0], [50, 72]])
# verification_data
'''2020.12.21
这里一定要注意使用float，不然用整数的话kriging工作会出问题
'''
load1, load2, load3 = 15.6, 28.6, 36.5
angle1, angle2, angle3 = 5.6, 36.6, 69.5

l123 = [load1, load2, load3]
a123 = [angle1, angle2, angle3]
load_list, angle_list = [], []
load_std, angle_std = [], []
for i in range(3):
    _load = np.loadtxt(load_np_txt_path + r"load\load_" + str(l123[i]) + ".txt")
    print("荷重传感器静止时均值与标准差：", _load.flatten().mean(), _load.flatten().std())
    load_std.append(_load.flatten().std())
    load_list.append(_load)
    _angle = np.loadtxt(load_np_txt_path + r"angle\angle_" + str(a123[i]) + ".txt")
    print("角度传感器静止时均值与标准差：", _angle.flatten().mean(), _angle.flatten().std())
    angle_std.append(_angle.flatten().std())
    angle_list.append(_angle)
print("荷重传感器静止时2*sigma：", np.mean(load_std) * 2)
print("角度传感器静止时2*sigma：", np.mean(angle_std) * 2)
independent_variables = []
for i in range(len(angle_list)):
    _temp = []
    for j in range(len(angle_list[0])):
        _temp_inner = []
        for h in range(len(angle_list[0][0])):
            _temp_inner.append([load_list[i][j][h], angle_list[i][j][h]])
        _temp.append(_temp_inner)
    independent_variables.append(_temp)
independent_variables = np.array(independent_variables)
# print(independent_variables)
train_low = create_train_independent_variables(force_arr_low, degree_arr_low)
multirbf_list = create_mf_rbf(train_low, train_high, array_real_stress_low, array_real_stress_high)

list_test_predict_stress_multi = []
for i in range(len(multirbf_list)):
    _temp_multi = []
    for j in range(len(independent_variables)):
        _temp_inner_ = []
        for h in range(len(independent_variables[0])):
            _temp_inner_.append(multirbf_list[i].predict(independent_variables[j][h]))
        _temp_multi.append(_temp_inner_)
    list_test_predict_stress_multi.append(_temp_multi)
# print(list_test_predict_stress_multi)
test_predict_stress_multi = np.array(list_test_predict_stress_multi)
print(test_predict_stress_multi.shape)
fig = plt.figure(figsize=(12, 6), dpi=100)
ax = fig.add_subplot(111)
axins = ax.inset_axes((0.35, 0.1, 0.3, 0.1))
xlim0, xlim1, ylim0, ylim1 = 40, 60, 15.5, 15.7
axins.set_xlim(xlim0, xlim1)
axins.set_ylim(ylim0, ylim1)
tx0 = xlim0
tx1 = xlim1
ty0 = ylim0
ty1 = ylim1
sx = [tx0, tx1, tx1, tx0, tx0]
sy = [ty0, ty0, ty1, ty1, ty0]
# 画两条线
"""
axesA是局部放大图的坐标轴
axesB是全局的坐标轴
"""
xy = (xlim0, ylim0)
xy2 = (xlim0, ylim1)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins, axesB=ax)
axins.add_artist(con)
xy = (xlim1, ylim0)
xy2 = (xlim1, ylim1)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins, axesB=ax)
axins.add_artist(con)

axins2 = ax.inset_axes((0.35, 0.4, 0.3, 0.1))
xlim0, xlim1, ylim0, ylim1 = 40, 60, 28.5, 28.7
axins2.set_xlim(xlim0, xlim1)
axins2.set_ylim(ylim0, ylim1)
tx0 = xlim0
tx1 = xlim1
ty0 = ylim0
ty1 = ylim1
sx2 = [tx0, tx1, tx1, tx0, tx0]
sy2 = [ty0, ty0, ty1, ty1, ty0]
# 画两条线
xy = (xlim0, ylim0)
xy2 = (xlim0, ylim1)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins2, axesB=ax)
axins2.add_artist(con)
xy = (xlim1, ylim0)
xy2 = (xlim1, ylim1)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins2, axesB=ax)
axins2.add_artist(con)

axins3 = ax.inset_axes((0.35, 0.8, 0.3, 0.1))
xlim0, xlim1, ylim0, ylim1 = 40, 60, 36.4, 36.6
axins3.set_xlim(xlim0, xlim1)
axins3.set_ylim(ylim0, ylim1)
tx0 = xlim0
tx1 = xlim1
ty0 = ylim0
ty1 = ylim1
sx3 = [tx0, tx1, tx1, tx0, tx0]
sy3 = [ty0, ty0, ty1, ty1, ty0]
# 画两条线
xy = (xlim0, ylim1)
xy2 = (xlim0, ylim0)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins3, axesB=ax)
axins3.add_artist(con)
xy = (xlim1, ylim1)
xy2 = (xlim1, ylim0)
con = ConnectionPatch(xyA=xy2, xyB=xy, coordsA="data", coordsB="data",
                      axesA=axins3, axesB=ax)
axins3.add_artist(con)

ax.set_title('The signal of load cell', fontsize=24)
# 角度
len_ = 540
load1, load2, load3 = 15.6, 28.6, 36.5
save_load_data1, save_load_data2, save_load_data3 = [], [], []
path_test = r"D:\Alai\paper_Alai\Papers_v1\1\v2\Papers\paper_result_truss\20200713_stress_data\\"
load = "25"  # 10 15 20 23 23_2 25_1 25 30 35 40 实际的重物质量(kg) 第一篇论文23_2 第二篇论文25,35
row_index = 3  # 因为应变采集仪在开始测量的时候总是慢一拍，所以要调整到和仿真的一致，从3开始调
x_tick = 105  # x的最大值
list_real_time_simulation_data = Read_SData(path_test + load + r".txt")
list_test_data = Read_TData(path_test + load + r".xlsx", row_index)
_num = min(len(list_real_time_simulation_data[0]), len(list_test_data[1][0]))
for i in range(10):
    load_temp = test_predict_stress_multi[0][1][i]
    # load_temp2 = test_predict_stress_multi[1]
    # load_temp3 = test_predict_stress_multi[2]
    save_load_data1.append(load_temp)
    # save_load_data2.append(load_temp2)
    # save_load_data3.append(load_temp3)
    ax.plot(list_test_data[0][0:_num],
            load_temp,
            lw=0.5,
            color='#ff7f0e',
            linestyle='-')
    # ax.plot(list_test_data[0][0:_num],
    #         load_temp2,
    #         lw=0.5,
    #         color='#1f77b4',
    #         linestyle='-')
    # ax.plot(list_test_data[0][0:_num],
    #         load_temp3,
    #         lw=0.5,
    #         color='#ff77f4',
    #         linestyle='-')
    axins.plot(list_test_data[0][0:_num],
               load_temp,
               lw=0.5,
               color='#ff7f0e',
               linestyle='-')
    # axins2.plot(list_test_data[0][0:_num],
    #             load_temp2,
    #             color='#1f77b4',
    #             lw=0.5,
    #             linestyle='-')
    # axins3.plot(list_test_data[0][0:_num],
    #             load_temp3,
    #             lw=0.5,
    #             color='#ff77f4',
    #             linestyle='-')
ax.plot(sx2, sy2, "black")
ax.plot(sx, sy, "black")
ax.plot(sx3, sy3, "black")
BIGGER_SIZE = 18
ax.set_xlabel("Time (s)", fontsize=BIGGER_SIZE)
ax.set_ylabel("Mass of load (Kg)", fontsize=BIGGER_SIZE)
ax.set_xlim(0, x_tick)
ax.set_ylim(0, 21)
x_ticks = np.arange(0, x_tick, 20)
y_ticks = np.arange(0, 21, 5)
ax.tick_params(labelsize=BIGGER_SIZE)
ax.set_xticks(x_ticks)
ax.set_yticks(y_ticks)
# ax.set_yticks(y_ticks2)
# 指定 data  设置的bottom(也就是指定的x轴)绑定到y轴的0这个点上
ax.spines['bottom'].set_position(('data', 0))
ax.spines['left'].set_position(('data', 0))
ax.grid()  # x坐标轴的网格使用主刻度
# plt.show()
for _obj in test_predict_stress_multi:
    _temp = []
    for _obj1 in _obj:
        temp = _obj1.flatten()
        _temp.append(temp.std())
        # print("应力的均值和标准差：", temp.mean(), temp.std())
    print("应力2*sigma：", np.mean(_temp) * 2)

# def save_to_excel():
#     pd_results_excel = pd.DataFrame(list_results_excel)
#     pd_results_excel.columns = ['verification_high', 'verification_low', 'verification_co', 'improvement']
#     pd_results_excel.index = range(1, 19)
#     writer = pd.ExcelWriter('r2_results_20210311_stress_gs.xlsx')  # 创建名称为hhh的excel表格
#     pd_results_excel.to_excel(writer, 'page_1',
#                               float_format='%.10f')  # float_format 控制精度，将data_df写到hhh表格的第一页中。若多个文件，可以在page_2中写入
#     writer.save()
